from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


class SpreadsheetResultRenderer:
    CHART_WIDTH = 1000
    CHART_HEIGHT = 600
    MAX_CHART_ITEMS = 20

    def to_xlsx(self, rows: list[dict[str, Any]], fallback_columns: list[str], plan: dict) -> bytes:
        columns = list(rows[0]) if rows else list(fallback_columns)
        numeric_columns = {item["alias"] for item in plan.get("aggregations", []) if item.get("op") != "count"}
        count_columns = {item["alias"] for item in plan.get("aggregations", []) if item.get("op") == "count"}
        workbook = Workbook(write_only=False)
        sheet = workbook.active
        sheet.title = "Result"
        sheet.freeze_panes = "A2"
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for column_index, name in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=column_index, value=name)
            cell.fill = header_fill
            cell.font = header_font
        for row_index, row in enumerate(rows, start=2):
            for column_index, column in enumerate(columns, start=1):
                value = row.get(column, "")
                if column in count_columns:
                    try:
                        value = int(value)
                    except (TypeError, ValueError):
                        pass
                elif column in numeric_columns:
                    try:
                        value = float(Decimal(str(value)))
                    except (InvalidOperation, TypeError, ValueError):
                        pass
                else:
                    value = self._safe_spreadsheet_text(value)
                sheet.cell(row=row_index, column=column_index, value=value)
        if columns:
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows) + 1)}"
        for index, column in enumerate(columns, start=1):
            values = [str(column), *(str(row.get(column, "")) for row in rows[:200])]
            sheet.column_dimensions[get_column_letter(index)].width = min(50, max(10, max(len(value) for value in values) + 2))
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def chart(self, rows: list[dict[str, Any]], plan: dict) -> tuple[bytes, dict] | None:
        group_by = list(plan.get("groupBy") or [])
        aggregations = list(plan.get("aggregations") or [])
        if not rows or not group_by or not aggregations:
            return None
        value_item = next((item for item in aggregations if item.get("op") in {"count", "sum", "avg", "min", "max"}), None)
        if value_item is None:
            return None
        label_column = group_by[0]
        value_column = value_item["alias"]
        values = []
        for row in rows[: self.MAX_CHART_ITEMS]:
            try:
                number = float(Decimal(str(row.get(value_column, ""))))
            except (InvalidOperation, TypeError, ValueError):
                return None
            values.append((str(row.get(label_column, "")), number))
        if not values:
            return None
        return self._bar_chart(values, value_column), {
            "chartType": "bar",
            "labelColumn": label_column,
            "valueColumn": value_column,
            "itemCount": len(values),
            "truncated": len(rows) > self.MAX_CHART_ITEMS,
        }

    def _bar_chart(self, values: list[tuple[str, float]], value_column: str) -> bytes:
        image = Image.new("RGB", (self.CHART_WIDTH, self.CHART_HEIGHT), "#FFFFFF")
        draw = ImageDraw.Draw(image)
        font = ImageFont.load_default()
        left, right, top, bottom = 90, 40, 60, 90
        plot_width = self.CHART_WIDTH - left - right
        plot_height = self.CHART_HEIGHT - top - bottom
        minimum = min(0.0, *(value for _, value in values))
        maximum = max(0.0, *(value for _, value in values))
        span = maximum - minimum or 1.0
        zero_y = top + int((maximum / span) * plot_height)
        draw.text((left, 20), f"Analysis result: {value_column}"[:100], fill="#17212B", font=font)
        draw.line((left, zero_y, left + plot_width, zero_y), fill="#56616B", width=2)
        slot = plot_width / max(1, len(values))
        bar_width = max(4, int(slot * 0.62))
        for index, (label, value) in enumerate(values):
            x0 = left + int(index * slot + (slot - bar_width) / 2)
            x1 = x0 + bar_width
            value_y = top + int(((maximum - value) / span) * plot_height)
            y0, y1 = sorted((zero_y, value_y))
            if y0 == y1:
                y0 = max(top, y0 - 1)
            color = "#2E75B6" if value >= 0 else "#C84A4A"
            draw.rectangle((x0, y0, x1, y1), fill=color)
            safe_label = label.encode("ascii", "replace").decode("ascii")[:14]
            draw.text((x0, self.CHART_HEIGHT - bottom + 14), safe_label, fill="#333333", font=font)
            draw.text((x0, max(top, y0 - 14)), f"{value:g}"[:14], fill="#333333", font=font)
        output = io.BytesIO()
        image.save(output, format="PNG", optimize=True)
        return output.getvalue()

    @staticmethod
    def _safe_spreadsheet_text(value: Any):
        if value is None:
            return ""
        if not isinstance(value, str):
            return value
        return "'" + value if value.startswith(("=", "+", "-", "@")) else value
